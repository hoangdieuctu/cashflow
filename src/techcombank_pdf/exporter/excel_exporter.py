"""Export transactions to a formatted Excel workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

from techcombank_pdf.models.transaction import ParseResult, TransactionType


def export_excel(result: ParseResult, output_path: str | Path) -> Path:
    """Export ParseResult to a formatted Excel workbook.

    Creates two sheets:
    - Transactions: detailed transaction list
    - Summary: totals and metadata
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Transactions sheet ---
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "Transaction Date",
        "Posting Date",
        "Description",
        "Original Amount",
        "Currency",
        "Billing Amount (VND)",
        "Type",
        "Category",
        "Merchant",
        "Card",
        "Reference",
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, txn in enumerate(result.transactions, 2):
        ws.cell(row=row_idx, column=1, value=txn.transaction_date)
        ws.cell(row=row_idx, column=2, value=txn.posting_date)
        ws.cell(row=row_idx, column=3, value=txn.description)
        ws.cell(row=row_idx, column=4, value=float(txn.original_amount))
        ws.cell(row=row_idx, column=5, value=txn.original_currency)
        ws.cell(row=row_idx, column=6, value=float(txn.billing_amount_vnd))
        ws.cell(row=row_idx, column=7, value=txn.transaction_type.value)
        ws.cell(row=row_idx, column=8, value=txn.category or "")
        ws.cell(row=row_idx, column=9, value=txn.merchant_name or "")
        ws.cell(row=row_idx, column=10, value=txn.card_last_four or "")
        ws.cell(row=row_idx, column=11, value=txn.reference_number or "")

    # Format date and number columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.number_format = "DD/MM/YYYY"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = "#,##0"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = "#,##0"

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(headers[col_idx - 1])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Summary sheet ---
    ws_summary = wb.create_sheet("Summary")
    summary_data = [
        ("Source File", result.metadata.source_file or ""),
        ("Statement Date", result.metadata.statement_date),
        ("Due Date", result.metadata.due_date),
        ("Card Number", result.metadata.card_number_masked or ""),
        ("", ""),
        ("Total Transactions", result.transaction_count),
        ("Total Debit", float(result.total_debit)),
        ("Total Credit", float(result.total_credit)),
        ("Net Amount", float(result.total_debit - result.total_credit)),
        ("", ""),
        ("Parse Method", result.parse_method),
        ("Pages", result.page_count),
    ]

    label_font = Font(bold=True)
    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_label = ws_summary.cell(row=row_idx, column=1, value=label)
        cell_label.font = label_font
        ws_summary.cell(row=row_idx, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30

    wb.save(str(output_path))
    return output_path
